import random
import string
import pandas as pd
import openpyxl
from datetime import datetime, timedelta

# 生成随机身份证号
def generate_ssn():
    return ''.join(random.choices(string.digits, k=13))

# 生成随机住址
def generate_address():
    return 'Address' + str(random.randint(1, 10))

# 生成随机工资
def generate_salary():
    return random.randint(1000, 5000)

# 生成随机部门号
def generate_dno():
    return '{:03d}'.format(random.randint(1, 5))

# 创建员工数据
employees = []
for i in range(1, 51):
    ename = 'E' + str(i)
    essn = generate_ssn()
    address = generate_address()
    salary = generate_salary()
    superssn = generate_ssn()
    dno = generate_dno()
    employees.append({'ENAME': ename, 'ESSN': essn, 'ADDRESS': address, 'SALARY': salary, 'SUPERSSN': superssn, 'DNO': dno})


# 生成随机开始工作日期
def generate_start_date():
    start_date = datetime(2004, 1, 1)
    end_date = datetime(2023, 12, 31)
    random_date = start_date + (end_date - start_date) * random.random()
    return random_date.strftime('%Y-%m-%d')

# 创建部门数据
departments = []
for i in range(1, 6):
    dname = 'D' + str(i)
    dnember = '{:03d}'.format(i)
    mgrssn = generate_ssn()
    mgrstartdate = generate_start_date()
    departments.append({'DNAME': dname, 'DNEMBER': dnember, 'MGRSSN': mgrssn, 'MGRSTARTDATE': mgrstartdate})
    
# 创建工程数据
projects = []
for i in range(1, 11):
    pno = 'P' + str(i)
    pname = 100 + i
    plocation = 'Address' + str(i)
    dno = '{:03d}'.format(random.randint(1, 5))
    projects.append({'PNAME': pname, 'PNO': pno, 'PLOCATION': plocation, 'DNO': dno})


# 创建工作数据
work_records = []
for i in range(1, 51):
    essn = '{:03d}'.format(i)
    pno = random.randint(100, 110)
    hours = random.randint(0, 10)
    work_records.append({'ESSN': essn, 'PNO': pno, 'HOURS': hours})
    


writer=pd.ExcelWriter('Learning/Database_System/homework/homework2/data.xlsx')


df = pd.DataFrame(employees)
df.to_excel(writer, sheet_name='E', index=False)
df = pd.DataFrame(departments)
df.to_excel(writer, sheet_name='D', index=False)
df = pd.DataFrame(projects)
df.to_excel(writer, sheet_name='P', index=False)
df = pd.DataFrame(work_records)
df.to_excel(writer, sheet_name='W', index=False)

writer.close()

#修改表数据使其符合实际

# 打开Excel文件
workbook = openpyxl.load_workbook('Learning/Database_System/homework/homework2/data.xlsx')

# 获取sheet1和sheet2
sheet1 = workbook['E']
sheet2 = workbook['W']

# 遍历sheet1中的每一行，将ESSN列的数据复制到sheet2中对应行的ESSN列
for row in range(2, sheet1.max_row + 1):
    essn_value = sheet1.cell(row=row, column=2).value  # 获取sheet1中ESSN列的值
    sheet2.cell(row=row, column=1, value=essn_value)   # type: ignore # 将该值写入sheet2中对应行的ESSN列

# 获取sheet1和sheet2
sheet2 = workbook['E']
sheet1 = workbook['D']

# 遍历sheet2中的每一行，根据DNO列的值在SUPERSSN列填入在sheet1中DNEMBER列取值相同行的MGRSSN列的值
for row in range(2, sheet2.max_row + 1):
    dno_value = sheet2.cell(row=row, column=6).value  # 获取sheet2中DNO列的值
    for row_sheet1 in range(2, sheet1.max_row + 1):
        if sheet1.cell(row=row_sheet1, column=2).value == dno_value:  # 查找sheet1中DNEMBER列取值与DNO相同的行
            mgrssn_value = sheet1.cell(row=row_sheet1, column=3).value  # 获取sheet1中DNEMBER列对应行的MGRSSN列的值
            sheet2.cell(row=row, column=5, value=mgrssn_value)   # type: ignore # 将该值写入sheet2中对应行的SUPERSSN列
            break

# 保存修改后的Excel文件
workbook.save('Learning/Database_System/homework/homework2/data.xlsx')